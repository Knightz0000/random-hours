import openpyxl
import datetime
import random

path = '/home/knightz____/workspace/projects/random-hours/planilhas/Rel_atividades 08-2024 KevinSilva.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

for i in range(10, 63):
    if i % 2 == 0:
        hour = random.randint(7, 8)
        hour_later = random.randint(11, 12)
        if hour == 7:
            minute = random.randint(56, 59)
        else:
            minute = random.randint(00, 5)
        if hour_later == 11:
            minute_later = random.randint(55, 59)
        else:
            minute_later = random.randint(00, 5)

        ws[f'E{i}'] = datetime.time(hour=hour, minute=minute)
        ws[f'E{i}'].number_format = 'H:MM'
        ws[f'F{i}'] = datetime.time(hour=hour_later, minute=minute_later)
        ws[f'F{i}'].number_format = 'H:MM'
    else:
        hour = random.randint(13, 14)
        hour_later = random.randint(17, 18)
        if hour == 13:
            minute = random.randint(55, 59)
        else:
            minute = random.randint(00, 15)
        if hour_later == 17:
            minute_later = random.randint(55, 59)
        else:
            minute_later = random.randint(00, 15)

        ws[f'E{i}'] = datetime.time(hour=hour, minute=minute)
        ws[f'E{i}'].number_format = 'H:MM'
        ws[f'F{i}'] = datetime.time(hour=hour_later, minute=minute_later)
        ws[f'F{i}'].number_format = 'H:MM'


wb.save(path)
